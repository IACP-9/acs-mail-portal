import os
import re
import json
import base64
from io import BytesIO

from flask import Flask, render_template, request
from azure.communication.email import EmailClient
from azure.identity import DefaultAzureCredential
from openpyxl import load_workbook

app = Flask(__name__)

ACS_ENDPOINT = os.getenv("ACS_ENDPOINT")
ACS_SENDER_ADDRESS = os.getenv("ACS_SENDER_ADDRESS")


def get_user_info(req):
    principal = req.headers.get("X-MS-CLIENT-PRINCIPAL")

    if not principal:
        return None, None

    decoded = base64.b64decode(principal)
    user_data = json.loads(decoded)

    name = None
    email = None

    for claim in user_data.get("claims", []):
        claim_type = claim.get("typ")
        claim_value = claim.get("val")

        if claim_type == "name":
            name = claim_value

        if claim_type in ["preferred_username", "emails"]:
            email = claim_value

    return name, email


def is_valid_email(email):
    pattern = r"^[^@\s]+@[^@\s]+\.[^@\s]+$"
    return re.match(pattern, email) is not None


def parse_manual_emails(raw_text):
    if not raw_text:
        return []

    normalized = raw_text.replace("\n", ",").replace(";", ",")
    parts = [item.strip() for item in normalized.split(",")]
    return [item for item in parts if item]


def parse_excel_emails(file_storage):
    emails = []

    if not file_storage or not file_storage.filename:
        return emails

    file_bytes = file_storage.read()
    if not file_bytes:
        return emails

    workbook = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    sheet = workbook.active

    headers = []
    for cell in sheet[1]:
        headers.append(str(cell.value).strip().lower() if cell.value else "")

    email_col_index = None
    for idx, header in enumerate(headers):
        if header in ["email", "correo", "mail", "correo electronico", "correo electrónico"]:
            email_col_index = idx
            break

    if email_col_index is None:
        raise ValueError("El archivo Excel debe tener una columna llamada 'email' o 'correo'.")

    for row in sheet.iter_rows(min_row=2, values_only=True):
        value = row[email_col_index] if email_col_index < len(row) else None
        if value:
            emails.append(str(value).strip())

    return emails


def clean_and_validate_emails(emails):
    unique_emails = []
    invalid_emails = []
    seen = set()

    for email in emails:
        email_clean = email.strip().lower()

        if not email_clean:
            continue

        if email_clean in seen:
            continue

        if is_valid_email(email_clean):
            unique_emails.append(email_clean)
            seen.add(email_clean)
        else:
            invalid_emails.append(email)

    return unique_emails, invalid_emails


def send_email(to_email, subject, body):
    credential = DefaultAzureCredential()
    client = EmailClient(ACS_ENDPOINT, credential)

    message = {
        "senderAddress": ACS_SENDER_ADDRESS,
        "recipients": {
            "to": [{"address": to_email}]
        },
        "content": {
            "subject": subject,
            "plainText": body,
            "html": f"""
            <html>
              <body>
                <p>{body}</p>
              </body>
            </html>
            """
        }
    }

    poller = client.begin_send(message)
    return poller.result()


@app.route("/", methods=["GET", "POST"])
def index():
    success = None
    error = None
    invalid_emails = []
    tracking_ids = []
    total_sent = 0

    user_name, user_email = get_user_info(request)

    if request.method == "POST":
        recipients_raw = request.form.get("recipients", "").strip()
        subject = request.form.get("subject", "").strip()
        body = request.form.get("body", "").strip()
        excel_file = request.files.get("excel_file")

        if not subject or not body:
            error = "Debes completar el asunto y el mensaje."
        else:
            try:
                manual_emails = parse_manual_emails(recipients_raw)
                excel_emails = parse_excel_emails(excel_file)

                all_emails = manual_emails + excel_emails
                valid_emails, invalid_emails = clean_and_validate_emails(all_emails)

                if not valid_emails:
                    error = "No se encontraron correos válidos para enviar."
                else:
                    for email in valid_emails:
                        result = send_email(email, subject, body)
                        tracking_ids.append({
                            "email": email,
                            "id": result.get("id")
                        })

                    total_sent = len(valid_emails)
                    success = f"Se enviaron {total_sent} correos correctamente."

            except Exception as ex:
                error = f"Ocurrió un error: {str(ex)}"

    return render_template(
        "index.html",
        success=success,
        error=error,
        invalid_emails=invalid_emails,
        tracking_ids=tracking_ids,
        total_sent=total_sent,
        sender_address=ACS_SENDER_ADDRESS,
        user_name=user_name,
        user_email=user_email
    )


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=8000)
